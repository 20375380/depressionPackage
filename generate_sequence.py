import cv2
import os
#import random
import torch
import numpy as np
from PIL import Image
#import efficientnet_pytorch
#import pandas as pd
#import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
from torchvision.transforms import transforms


#datasets_root = r'D:\\experiment\\picture'
#datasets_new_root = r'D:\\experiment\\fuse'
#data_update = r'D:\try\dataset'

class seq_generator():
    def __init__(self,video_path,datasets_root,datasets_new_root,txt8_path,excel):
        self.video_path = video_path # 带处理视频路径
        self.datasets_root = datasets_root # 剪切好的帧图像保存路径：路径中不能含有中文 以下同上
        self.datasets_new_root = datasets_new_root # 经过多时相融合后的图像保存路径
        self.txt8_path = txt8_path
        self.excel = excel
        self.model = 0
        self.j = 0
        self.serial_num = 0

    '''
    此函数用来将视频以每秒24帧的频率剪切成图片
    '''
    def cut_frame(self):

        cap = cv2.VideoCapture(self.video_path)  # 存放待预测视频的路径（改）;
        fourcc = cv2.VideoWriter_fourcc(*'XVID')
        fps = cap.get(cv2.CAP_PROP_FPS)
        # fps=30
        size = (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)))
        # size=(960,546)
        i = 0
        while (cap.isOpened()):
            i = i + 1
            ret, frame = cap.read()
            if ret == True:
                cv2.imwrite(self.datasets_root + '//' + str(i) + '.jpg', frame)  # 将要保存一帧帧图片的路径;
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
            else:
                break
        cap.release()
        cv2.destroyAllWindows()
        print('cut frame successful')

    '''
     此函数用来将剪切好的帧图片进行多时相图像融合处理
    '''
    def pretreatment(self):

        all_folds = len(os.listdir(self.datasets_root))+1
        for each_frame in range(33,all_folds):
            c,d = divmod(each_frame,16)
            if d >= 8:
                img2 = cv2.imread(self.datasets_root + '/'  + str(int(each_frame)) + '.jpg')
                img1 = cv2.imread(self.datasets_root + '/'  + str(int(each_frame) - 8) + '.jpg')
                image = cv2.absdiff(img2,img1)
                image1 = cv2.absdiff(img2,img1)
            #all_folds_num = len(os.listdir(datasets_new_root))
            #image = cv2.imread(path_read)
            #image1 = cv2.imread(path_read)
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

                gradX = cv2.Sobel(gray, ddepth=cv2.CV_32F, dx=1, dy=0, ksize=-1)
                gradY = cv2.Sobel(gray, ddepth=cv2.CV_32F, dx=0, dy=1, ksize=-1)
                gradient = cv2.subtract(gradX, gradY)
                gradient = cv2.convertScaleAbs(gradient)

                blurred = cv2.blur(gradient, (13,13))
                (_, thresh) = cv2.threshold(blurred, 90, 255, cv2.THRESH_BINARY)

                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 25))
                closed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)

                closed = cv2.erode(closed, None, iterations=1)
                closed = cv2.dilate(closed, None, iterations=1)

                (cnts, _) = cv2.findContours(closed.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                c = sorted(cnts, key=cv2.contourArea, reverse=True)[0]
            # compute the rotated bounding box of the largest contour
                rect = cv2.minAreaRect(c)
                box = np.int0(cv2.boxPoints(rect))
            # draw a bounding box arounded the detected barcode and display the image
                cv2.drawContours(image, [box], -1, (0, 255, 0), 3)

                Xs = [i[0] for i in box]
                Ys = [i[1] for i in box]
                x1 = min(Xs)
                x2 = max(Xs)
                y1 = min(Ys)
                y2 = max(Ys)
                hight = y2 - y1
                width = x2 - x1
                cropImg = image1[y1:y1+hight, x1:x1+width]
            #img = cv2.imread(cropImg)
            #datasets_new_root = r'D:\excute\a'
                imgSize = cropImg.shape[0:2]
                save_num = len(os.listdir(self.datasets_new_root)) + 1
                if imgSize[0] < 150 or imgSize[1] < 150:
                    cv2.imwrite(self.datasets_new_root + '//'+str(save_num)+ '.jpg', image1)
                else:
                    cv2.imwrite(self.datasets_new_root + '//'+str(save_num)+ '.jpg', cropImg)
            #cv2.imwrite(datasets_new_root + '//' + str(int(all_folds_num)+1) + '.jpg',a1)
        print('fuse successful')

    def predict(self,path):
        img = Image.open(path)
        #img1 = cv2.imread(path)
        val_transforms = transforms.Compose([transforms.Resize(224),
                                         transforms.CenterCrop(224),
                                         transforms.ToTensor(),
                                         transforms.Normalize([0.485, 0.456, 0.406], [0.249, 0.245, 0.245])
                                         ])

        img_tensor = val_transforms(img)
        img_tensor = img_tensor.unsqueeze(dim=0)
        #model.to('cpu')
        res = self.model(img_tensor)
        idx = torch.argmax(res).item()
        #save_num_new = len(os.listdir(data_update + '//'+ str(idx))) + 1
        #cv2.imwrite(data_update + '//'+str(idx) + '//'+ str(save_num_new)+ '.jpg', img1)

        #print ('Class: ',idx)
        with open (self.txt8_path,'a',encoding='utf-8') as f:  #存放单帧预测结果的路径（改）：和下面的相同 TXT文本我是以日期_哪只老鼠_哪个摄像头_8(代表一个序列16张图像融合成8张)例如：2020.11.15-46-ch02-8.txt
            f.write(str(idx))
            f.write(' ')
            self.j = self.j + 1
            a,b = divmod(self.j, 8)
            if b == 0:
                f.write('\n')
            else:
                pass
        #return idx


        '''
        批量预测
        '''
    def batch_excute(self,root_dir):

        b = os.listdir(root_dir)
        #print(a)
        for i in range(1,len(b)):
            path = root_dir + '/'  + str(i) + '.jpg'
    #       c,d = divmod(i,16)
    #       if d < 8:
            self.predict(path)
        print('finish batch_execute')


    '''
    此函数用于预测多时相图像的类别
    '''
    def classify_main(self):
        print('analysing')
        self.model = torch.load('model_b4_8frame_11_cpu.pkl') # 读取模型,map_location=torch.device('cuda')
        print('load model successful')
        root_dir = self.datasets_new_root
        path = self.batch_excute(root_dir)
        print('classify successful')
        # predict(path)



    '''
    此函数用于最终对16帧图像进行整体的预测
    '''
    def last_predict(self):
        self.serial_num = 1
        list = []
        line_11 = []
        result = []
        dict1 = []
        dict2 = []
        b = {}
        for line in open(self.txt8_path):   #单张预测文本（改）同上 两个名字保持一致,保存按照10换行的数据的txt
            list.append(line)
        # txt-1中保存的是每一行中出现次数最多的
        for i in range(len(list)):
            line_11 = list[i].split(' ')[0:-1]
            result = max(set(line_11), key=line_11.count)
            dict1.append(result)
        print(dict1)
        for i in dict1:
            if dict1.count(i) > 1:
                b[i] = dict1.count(i)
        print(b)

        for self.j in range(len(dict1)):
            if dict1[self.j] == '0':
                dict2.append('groomhead')

            elif dict1[self.j] == '1':
                dict2.append('lickforepaws')

            elif dict1[self.j] == '2':
                dict2.append('lickabdomen')#舔腹

            elif dict1[self.j] == '3':
                dict2.append('lickback')

            elif dict1[self.j] == '4':
                dict2.append('licktail')

            elif dict1[self.j] == '5':
                dict2.append('laydown')

            elif dict1[self.j] == '6':
                dict2.append('rearpause')

            elif dict1[self.j] == '7':
                dict2.append('rear')

            elif dict1[self.j] == '8':
                dict2.append('climb')

            elif dict1[self.j] == '9':
                dict2.append('turn')

            elif dict1[self.j] == '10':
                dict2.append('walk')

        print(dict2)


        wb = Workbook()
        name = self.excel #之前杜晨师兄要的序列的表格保存路径及命名（改）
        wb.save(name)
        wb = load_workbook(name)
        sheet = wb.active
        sheet['A1'] = 'time'
        sheet['B1'] = 'event'
        sheet['A2'] = '0'
        sheet['B2'] = ':'

        for z in range(3,len(dict2)+3):
            a = 'A' + str(z)
            b = 'B' + str(z)
            sheet[a] = self.serial_num
            sheet[b] = dict2[self.serial_num-1]
            self.serial_num += 1

        c = 'A' + str(len(dict2)+3)
        d = 'B' + str(len(dict2)+3)
        sheet[c] = self.serial_num
        sheet[d] = '&'
        wb.save(name)
        print('last predict successful')


